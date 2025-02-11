import pandas as pd
import openpyxl
from openpyxl import Workbook
from datetime import datetime

# a mix of some things I have done before plus some new things helped with by a Code Assistant (GitHub COPilot)
# the NaN issues and some envoding issues were assisted much by the Code Assistant - the original document was output as export by Atlassian Jira

def create_issue_counts():
    # Setup filenames and read data
    filedate = datetime.now().strftime("%Y%m%d")
    output_filename = f'issue_counts_{filedate}.xlsx'

    try:
        df = pd.read_csv('requestTypes_Issues.csv', encoding='latin-1')
    except UnicodeDecodeError:
        df = pd.read_csv('requestTypes_Issues.csv', encoding='cp1252')

    # Get counts and types
    empty_request_types = len(df[df['Request Type'].isna()])
    request_types = sorted([rt for rt in df['Request Type'].unique() if pd.notna(rt)])

    # Initialize workbook
    wb = Workbook()
    wb.remove(wb.active)
    summary_sheet = wb.create_sheet('Summary')
    summary_sheet.append(['Request Type', 'Issue Type', 'Count', 'Other', 'Total Issues', 'Empty Request Types'])

    # Process data
    summary_rows = []
    grand_total = 0

    for req_type in request_types:
        req_type_df = df[df['Request Type'] == req_type]
        total_issues = 0
        other_count = len(req_type_df)
        issues_data = []

        # Process Operational Support Issues
        op_counts = req_type_df['ACCESS Operational Support Issues'].value_counts()
        for issue, count in op_counts.items():
            if pd.notna(issue):
                issues_data.append([req_type, issue, count, '', '', ''])
                total_issues += count
                other_count -= count

        # Process User Support Issues
        user_counts = req_type_df['ACCESS User Support Issue'].value_counts()
        for issue, count in user_counts.items():
            if pd.notna(issue):
                issues_data.append([req_type, issue, count, '', '', ''])
                total_issues += count
                other_count -= count

        # Sort and add data
        issues_data.sort(key=lambda x: str(x[1]))
        summary_rows.extend(issues_data)

        if other_count > 0:
            summary_rows.append([req_type, 'Uncategorized', '', other_count, '', ''])
            total_issues += other_count

        summary_rows.append([req_type, 'TOTAL', '', '', total_issues, ''])
        grand_total += total_issues

        # Create and populate individual sheet
        sheet_name = str(req_type)[:31]
        req_sheet = wb.create_sheet(sheet_name)
        req_sheet.append(['Issue Type', 'Count', 'Other', 'Total Issues', 'Empty Request Types'])

        for row in issues_data:
            req_sheet.append([row[1], row[2], '', '', ''])
        if other_count > 0:
            req_sheet.append(['Uncategorized', '', other_count, '', ''])
        req_sheet.append(['TOTAL', '', '', total_issues, ''])

    # Add all data to summary sheet
    for row in summary_rows:
        summary_sheet.append(row)
    summary_sheet.append(['GRAND TOTAL', '', '', '', grand_total, empty_request_types])

    # Save workbook
    wb.save(output_filename)

if __name__ == '__main__':
    create_issue_counts()
